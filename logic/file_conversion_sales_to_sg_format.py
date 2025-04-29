"""
CODE PURPOSE: To convert original file to SG file format. 
"""

# import module
import os
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import re

# logging setup
logging.basicConfig(
    level=logging.INFO, 
    format="%(asctime)s - %(levelname)s - %(message)s", 
    datefmt="%Y-%m-%d %H:%M:%S"
    
    )

def create_sg_format_table():
    sg_format = {

        'ConstID': [],
        'CustomerID': [],
        'CreditCardNo': [],
        'ExpiryDate': [],
        'CustomerName': [],
        'FirstName': [],
        'LastName': [],
        'NRIC': [],
        'NRIC_Old': [],
        'AccountNo': [],
        'NewAccountNo': [],
        'PolicyNo': [],
        'MarkForEnrol': [],
        'EnrolAction': [],
        'EnrolDate': [],
        'EnrolCode': [],
        'EnrolDescription': [],
        'Bank': [],
        'SourceCode': [],
        'CampaignCode': [],
        'SerialNo': [],
        'BatchNo': [],
        'Address1': [],
        'Address2': [],
        'Address3': [],
        'Address4': [],
        'Postcode': [],
        'City': [],
        'State': [],
        'TelHse': [],
        'TelOff': [],
        'TelHP': [],
        'FaxNumber': [],
        'Email': [],
        'SubDate': [],
        'CancelDate': [],
        'RejectDate': [],
        'ReactivateDate': [],
        'CardType': [],
        'IssuingBank': [],
        'Frequency': [],
        'DonationAmount': [],
        'TotalContribution': [],
        'DonorStatus': [],
        'AgentID': [],
        'Remarks': [],
        'DonorStatusRemarks': [],
        'RefNo': [],
        'LatestAnniversary': [],
        'LatestCollectedDate': [],
        'LatestCollectedAmount': [],
        'LastModified': [],
        'LastModifiedBy': [],
        'CreatedDate': [],
        'CreatedBy': [],
        'A0': [],
        'A1': [],
        'A2': [],
        'A3': [],
        'A4': [],
        'D0': [],
        'D1': [],
        'D2': [],
        'D3': [],
        'D4': [],
        'LastSent': [],
        'SourceCode2': [],
        'CycleDate': [],
        'AppcoStatus': [],
        'WeekEndingDate': [],
        'StatusDate': [],
        'AppcoBatchNo': [],
        'LatestCollectedStatus': [],
        'A0Attempts': [],
        'LatestNDNHFixDate': [],
        'LatestNDNHFixBy': [],
        'IsPendingBilling': [],
        'CancelCode': [],
        'CancelledBy': [],
        'DOB': [],
        'Gender': [],
        'ChildrenUnder18': [],
        'CancelSource': [],
        'ReactivateBy': [],
        'Race': [],
        'Title': [],
        'Event': [],
        'A0_Gross': [],
        'SourceCode1': [],
        'A0_Frequency': [],
        'CVV2': [],
        'optional_one': [],
        'optional_two': [],
        'optional_three': [],
        'Payment Category': [],
        'PolicyNoDate': [],
        'InstantDebit': [],
        'PLEDGETYPE': [],
        'DOBOTYPE': [],
        'PRINCIPAL': [],
        'OnBehalf_Title': [],
        'OnBehalf_FirstName': [],
        'OnBehalf_LastName': [],
        'OnBehalf_AlternateName': [],
        'OnBehalf_Add1': [],
        'OnBehalf_Add2': [],
        'OnBehalf_Add3': [],
        'OnBehalf_Add4': [],
        'OnBehalf_Postcode': [],
        'OnBehalf_City': [],
        'OnBehalf_State': []
    }

    return pd.DataFrame(sg_format)

def extract_data_from_salesfile_to_sgformatfile(new_df, df):
    new_df['ConstID'] = ''
    new_df['CustomerID'] = ''
    new_df['CreditCardNo'] = df['CREDIT CARD']
    new_df['ExpiryDate'] = df['EXPIRY']
    new_df['CustomerName'] = df['FIRSTNAME'] + df['LASTNAME']
    new_df['FirstName'] = df['FIRSTNAME']
    new_df['LastName'] = df['LASTNAME']
    new_df['NRIC'] = df['IC NUMBER']
    new_df['NRIC_Old'] = ''
    new_df['AccountNo'] = df['ACCOUNT NUMBER']
    new_df['NewAccountNo'] = ''
    new_df['PolicyNo'] = ''
    new_df['MarkForEnrol'] = df['ACCOUNT NUMBER'].apply(lambda x : 'false' if x == '' else 'TRUE')
    new_df['EnrolAction'] = ''
    new_df['EnrolDate'] = ''
    new_df['EnrolCode'] = ''
    new_df['EnrolDescription'] = ''
    new_df['Bank'] = df['PROCESSING BANK']
    new_df['SourceCode'] = df['RECRUITER CODE']
    new_df['CampaignCode'] = ''
    new_df['SerialNo'] = df['SERIAL NO']
    new_df['BatchNo'] = ''
    new_df['Address1'] = df['ADDRESS 1']
    new_df['Address2'] = df['ADDRESS 2']
    new_df['Address3'] = ''
    new_df['Address4'] = ''
    new_df['Postcode'] = df['POSTCODE']
    new_df['City'] = df['CITY']
    new_df['State'] = df['STATE']
    new_df['TelHse'] = ''
    new_df['TelOff'] = ''
    new_df['TelHP'] = df['TEL HP']
    new_df['FaxNumber'] = ''
    new_df['Email'] = df['EMAIL']
    new_df['SubDate'] = df['SIGNUP DATE']
    new_df['CancelDate'] = ''
    new_df['RejectDate'] = ''
    new_df['ReactivateDate'] = ''
    new_df['CardType'] = df['CARDTYPE']
    new_df['IssuingBank'] = df['ISSUING BANK']
    new_df['Frequency'] = df['FREQUENCY']
    new_df['DonationAmount'] = df['DONATION AMOUNT']
    new_df['TotalContribution'] = ''
    new_df['DonorStatus'] = 'A'
    new_df['AgentID'] = df['AGENT ID']
    new_df['Remarks'] = ''
    new_df['DonorStatusRemarks'] = ''
    new_df['RefNo'] = ''
    new_df['LatestAnniversary'] = ''
    new_df['LatestCollectedDate'] = df['SIGNUP DATE']
    new_df['LatestCollectedAmount'] = ''
    new_df['LastModified'] = ''
    new_df['LastModifiedBy'] = ''
    new_df['CreatedDate'] = df['SIGNUP DATE']
    new_df['CreatedBy'] = ''
    new_df['A0'] = ''
    new_df['A1'] = ''
    new_df['A2'] = ''
    new_df['A3'] = ''
    new_df['A4'] = ''
    new_df['D0'] = df['SIGNUP DATE']
    new_df['D1'] = ''
    new_df['D2'] = ''
    new_df['D3'] = ''
    new_df['D4'] = ''
    new_df['LastSent'] = ''
    new_df['SourceCode2'] = ''
    new_df['CycleDate'] = ''
    new_df['AppcoStatus'] = ''
    new_df['WeekEndingDate'] = ''
    new_df['StatusDate'] = ''
    new_df['AppcoBatchNo'] = ''
    new_df['LatestCollectedStatus'] = ''
    new_df['A0Attempts'] = ''
    new_df['LatestNDNHFixDate'] = ''
    new_df['LatestNDNHFixBy'] = ''
    new_df['IsPendingBilling'] = ''
    new_df['CancelCode'] = ''
    new_df['CancelledBy'] = ''
    new_df['DOB'] = df['DOB']
    new_df['Gender'] = df['GENDER']
    new_df['ChildrenUnder18'] = ''
    new_df['CancelSource'] = ''
    new_df['ReactivateBy'] = ''
    new_df['Race'] = df['RACE']
    new_df['Title'] = df['TITLE']
    new_df['Event'] = df['EVENT CODE']
    new_df['A0_Gross'] = ''
    new_df['SourceCode1'] = ''
    new_df['A0_Frequency'] = ''
    new_df['CVV2'] = ''
    new_df['optional_one'] = ''
    new_df['optional_two'] = ''
    new_df['optional_three'] = ''
    new_df['Payment Category'] = df['CARDTYPE'] + " " + df['DEBIT_CC_CARD']
    new_df['PolicyNoDate'] = ''
    new_df['InstantDebit'] = ''
    new_df['PLEDGETYPE'] = 'Standard'
    new_df['DOBOTYPE'] = ''
    new_df['PRINCIPAL'] = ''
    new_df['OnBehalf_Title'] = ''
    new_df['OnBehalf_FirstName'] = ''
    new_df['OnBehalf_LastName'] = ''
    new_df['OnBehalf_AlternateName'] = ''
    new_df['OnBehalf_Add1'] = ''
    new_df['OnBehalf_Add2'] = ''
    new_df['OnBehalf_Add3'] = ''
    new_df['OnBehalf_Add4'] = ''
    new_df['OnBehalf_Postcode'] = ''
    new_df['OnBehalf_City'] = ''
    new_df['OnBehalf_State'] = ''


    return new_df

def transform_card_type(x):
    if x == 'DEBIT CARD':
        return 'DC'
    elif x == 'CREDIT CARD':
        return 'CC'
    else:
        x

def create_new_file_name(file):
    match = re.search(r'\b\d{6}\b', file)
    return f'NewPledges - {match.group()}.xlsx'

def main():
    folder_path = r'C:\Users\mfmohammad\OneDrive - UNICEF\Documents\SW process\20250429'

    for file in os.listdir(folder_path):
        if 'ExportDailyDataSG' in file or 'ExportDailyData' in file:

            file_path = os.path.join(folder_path, file)

            logging.info('Read file')
            data = pd.read_excel(file_path, dtype={'EXPIRY' : str, 'ACCOUNT NUMBER' : str, 'POSTCODE' : str , 'TEL HP': str})

            data = data.fillna('')
            
            data['IC NUMBER'] = data['IC NUMBER'].where(data['IC NUMBER'] != '' , data['PASSPORT NUMBER'])

            logging.info('Convert card type')
            data['DEBIT_CC_CARD'] = data['DEBIT_CC_CARD'].apply(lambda x : transform_card_type(x))

            logging.info('Make data in title format')
            data['CARDTYPE'] = data['CARDTYPE'].str.title()

            logging.info('Replace blank in CARDTYPE with MBB')
            data['CARDTYPE'] = data['CARDTYPE'].apply(lambda x: 'MBB' if pd.isna(x) or x == '' else x)
            data['DEBIT_CC_CARD'] = data['DEBIT_CC_CARD'].fillna('')

            logging.info('Create SG format template')
            new_df = create_sg_format_table()

            logging.info('Copy data from sales file to template')
            new_df = extract_data_from_salesfile_to_sgformatfile(new_df, data)

            new_df['CreatedDate'] = pd.to_datetime(new_df['CreatedDate'], format='%d/%m/%Y').dt.strftime('%d-%m-%Y')
            new_df['D0'] = pd.to_datetime(new_df['D0'], format='%d/%m/%Y').dt.strftime('%d-%m-%Y')
            new_df['DOB'] = pd.to_datetime(new_df['DOB'], format='%d/%m/%Y').dt.strftime('%d-%m-%Y')
            new_df['Payment Category'] = new_df['Payment Category'].apply(lambda x: 'AMEX' if isinstance(x, str) and x.strip().lower() == 'amex cc' else x)
            
            output_file_path = os.path.join(folder_path, create_new_file_name(file))
            
            logging.info('Saving file...')
            new_df.to_excel(output_file_path, index=False)

            logging.info('Open saved file')
            wb = load_workbook(output_file_path)
            ws = wb.active

            # Step 3: Identify the column indices for "SubDate" and "LatestCollectedDate"
            subdate_col_idx = None
            latest_collected_col_idx = None

            logging.info('Reformat date columns')
            # Find the correct column indices
            for col_idx, col_name in enumerate(ws[1], start=1):
                if col_name.value == "SubDate":
                    subdate_col_idx = col_idx
                elif col_name.value == "LatestCollectedDate":
                    latest_collected_col_idx = col_idx

            # Create a date format style (check if style already exists to avoid duplicate name error)
            if "date_style" not in wb.named_styles:
                date_style = NamedStyle(name="date_style", number_format="DD-MM-YYYY")
                wb.add_named_style(date_style)
            else:
                date_style = wb.named_styles["date_style"]

            # Helper function to insert date formulas
            def apply_date_formula(column_index, data_series):
                if column_index:
                    for row_idx, d in enumerate(data_series, start=2):  # Start from row 2 (skip header)
                        if isinstance(d, str) and len(d) == 10:
                            formula = f"=DATE({d[-4:]},{d[3:5]},{d[:2]})"
                            cell = ws.cell(row=row_idx, column=column_index, value=formula)
                            cell.style = date_style

            # Apply to both columns
            apply_date_formula(subdate_col_idx, new_df["SubDate"])
            apply_date_formula(latest_collected_col_idx, new_df["LatestCollectedDate"])

            logging.info('Saving file')
            wb.save(output_file_path)

            logging.info('Process complete. File saved in folder.')

        

if __name__ == "__main__":
    main()