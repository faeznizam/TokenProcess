"""
CODE PURPOSE: 
1. To convert original file to SG file format. 
2. To extract data from result file (.csv) and reformat in table form in .xlsx
"""

# import module
import os
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import re
import time
import numpy as np


# logging setup
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s", datefmt="%Y-%m-%d %H:%M:%S")


# use vectorize dataframe instead of initalize empty dataframe method
def generate_sg_format_df(df):
    # generate sg format dataframe with salesfile data

    sg_df = pd.DataFrame({
        'ConstID': '',
        'CustomerID': '',
        'CreditCardNo': df['CREDIT CARD'],
        'ExpiryDate': df['EXPIRY'],
        'CustomerName': df['FIRSTNAME'] + df['LASTNAME'],
        'FirstName': df['FIRSTNAME'],
        'LastName': df['LASTNAME'],
        'NRIC': df['IC NUMBER'],
        'NRIC_Old': '',
        'AccountNo': df['ACCOUNT NUMBER'],
        'NewAccountNo': '',
        'PolicyNo': '',
        'MarkForEnrol': df['ACCOUNT NUMBER'].apply(lambda x: 'false' if x == '' else 'TRUE'),
        'EnrolAction': '',
        'EnrolDate': '',
        'EnrolCode': '',
        'EnrolDescription': '',
        'Bank': df['PROCESSING BANK'],
        'SourceCode': df['RECRUITER CODE'],
        'CampaignCode': '',
        'SerialNo': df['SERIAL NO'],
        'BatchNo': '',
        'Address1': df['ADDRESS 1'],
        'Address2': df['ADDRESS 2'],
        'Address3': '',
        'Address4': '',
        'Postcode': df['POSTCODE'],
        'City': df['CITY'],
        'State': df['STATE'],
        'TelHse': '',
        'TelOff': '',
        'TelHP': df['TEL HP'],
        'FaxNumber': '',
        'Email': df['EMAIL'],
        'SubDate': df['SIGNUP DATE'],
        'CancelDate': '',
        'RejectDate': '',
        'ReactivateDate': '',
        'CardType': df['CARDTYPE'],
        'IssuingBank': df['ISSUING BANK'],
        'Frequency': df['FREQUENCY'],
        'DonationAmount': df['DONATION AMOUNT'],
        'TotalContribution': '',
        'DonorStatus': 'A',
        'AgentID': df['AGENT ID'],
        'Remarks': '',
        'DonorStatusRemarks': '',
        'RefNo': '',
        'LatestAnniversary': '',
        'LatestCollectedDate': df['SIGNUP DATE'],
        'LatestCollectedAmount': '',
        'LastModified': '',
        'LastModifiedBy': '',
        'CreatedDate': df['SIGNUP DATE'],
        'CreatedBy': '',
        'A0': '',
        'A1': '',
        'A2': '',
        'A3': '',
        'A4': '',
        'D0': df['SIGNUP DATE'],
        'D1': '',
        'D2': '',
        'D3': '',
        'D4': '',
        'LastSent': '',
        'SourceCode2': '',
        'CycleDate': '',
        'AppcoStatus': '',
        'WeekEndingDate': '',
        'StatusDate': '',
        'AppcoBatchNo': '',
        'LatestCollectedStatus': '',
        'A0Attempts': '',
        'LatestNDNHFixDate': '',
        'LatestNDNHFixBy': '',
        'IsPendingBilling': '',
        'CancelCode': '',
        'CancelledBy': '',
        'DOB': df['DOB'],
        'Gender': df['GENDER'],
        'ChildrenUnder18': '',
        'CancelSource': '',
        'ReactivateBy': '',
        'Race': df['RACE'],
        'Title': df['TITLE'],
        'Event': df['EVENT CODE'],
        'A0_Gross': '',
        'SourceCode1': '',
        'A0_Frequency': '',
        'CVV2': '',
        'optional_one': '',
        'optional_two': '',
        'optional_three': '',
        'Payment Category': df['CARDTYPE'],
        'PolicyNoDate': '',
        'InstantDebit': '',
        'PLEDGETYPE': 'Standard',
        'DOBOTYPE': '',
        'PRINCIPAL': '',
        'OnBehalf_Title': '',
        'OnBehalf_FirstName': '',
        'OnBehalf_LastName': '',
        'OnBehalf_AlternateName': '',
        'OnBehalf_Add1': '',
        'OnBehalf_Add2': '',
        'OnBehalf_Add3': '',
        'OnBehalf_Add4': '',
        'OnBehalf_Postcode': '',
        'OnBehalf_City': '',
        'OnBehalf_State': ''
    })

    return sg_df

def transform_payment_category(x):
    
    if x == 'MBB':
        'MBB'

def transform_card_type(x):
    # using 
    lookup = {
        'debit card' : 'DC',
        'credit card' : 'CC'
    }

    if isinstance(x, str):
        return lookup.get(x.lower(), x)
    return x


def create_new_file_name(file):
    match = re.search(r'\b\d{6}\b', file)
    return f'NewPledges - {match.group()}.xlsx'


def convert_sw_file(folder_path, file):
    logging.info('Process 1: Create file with SG format using SW file')
    file_path = os.path.join(folder_path, file)

    logging.info(f'Read {file}')
    data = pd.read_excel(file_path, dtype={'EXPIRY' : str, 'ACCOUNT NUMBER' : str, 'POSTCODE' : str , 'TEL HP': str})
    
    logging.debug('Fill blank ic number in IC NUMBER column with data from PASSPORT NUMBER column')
    data['IC NUMBER'] = data['IC NUMBER'].fillna(data['PASSPORT NUMBER'])

    logging.debug('Transform card type according to its classification')
    data['DEBIT_CC_CARD'] = data['DEBIT_CC_CARD'].apply(lambda x : transform_card_type(x))

    logging.debug('Transform data in CARDTYPE in title case')
    data['CARDTYPE'] = data['CARDTYPE'].str.title()

    logging.debug('Replace blank in CARDTYPE with MBB marking and combine data with DEBIT_CC_CARD')
    data['CARDTYPE'] = data['CARDTYPE'].apply(lambda x: 'MBB' if pd.isna(x) or x == '' else x)
    data['CARDTYPE'] = np.where(data['CARDTYPE'] == 'MBB', data['CARDTYPE'], data['CARDTYPE'] + ' ' + data['DEBIT_CC_CARD'])
    
    logging.info('Create SG Format table with SW data')
    new_df = generate_sg_format_df(data)

    logging.debug('Reformat CreatedDate, D0, and DOB columns date to DD/MM/YYYY')
    new_df['CreatedDate'] = pd.to_datetime(new_df['CreatedDate'], format='%d/%m/%Y').dt.strftime('%d-%m-%Y')
    new_df['D0'] = pd.to_datetime(new_df['D0'], format='%d/%m/%Y').dt.strftime('%d-%m-%Y')
    new_df['DOB'] = pd.to_datetime(new_df['DOB'], format='%d/%m/%Y').dt.strftime('%d-%m-%Y')

    logging.debug("Replace 'AMEX' with 'amex cc' in Payment Category column")
    new_df['Payment Category'] = new_df['Payment Category'].apply(lambda x: 'AMEX' if isinstance(x, str) and x.strip().lower() == 'amex cc' else x)
    
    logging.debug(f'Create new file name: {create_new_file_name(file)}')
    file_name = create_new_file_name(file)

    logging.debug('Create file path to save result')
    output_file_path = os.path.join(folder_path, file_name)
    
    logging.info(f'Save {file_name} in {folder_path}')
    new_df.to_excel(output_file_path, index=False)
    
    logging.info(f'Re-open {file_name} file.')
    wb = load_workbook(output_file_path)
    ws = wb.active

    # Step 3: Identify the column indices for "SubDate" and "LatestCollectedDate"
    subdate_col_idx = None
    latest_collected_col_idx = None

    logging.info('Reformat date columns')
    # Find the correct column indices
    for col_idx, col_name in enumerate(ws[1], start=1):
        if col_name.value == "SubDate":
            subdate_col_idx = col_idx
        elif col_name.value == "LatestCollectedDate":
            latest_collected_col_idx = col_idx

    # Create a date format style (check if style already exists to avoid duplicate name error)
    if "date_style" not in wb.named_styles:
        date_style = NamedStyle(name="date_style", number_format="DD-MM-YYYY")
        wb.add_named_style(date_style)
    else:
        date_style = wb.named_styles["date_style"]

    # Helper function to insert date formulas
    def apply_date_formula(column_index, data_series):
        if column_index:
            for row_idx, d in enumerate(data_series, start=2):  # Start from row 2 (skip header)
                if isinstance(d, str) and len(d) == 10:
                    formula = f"=DATE({d[-4:]},{d[3:5]},{d[:2]})"
                    cell = ws.cell(row=row_idx, column=column_index, value=formula)
                    cell.style = date_style

    logging.debug('Reformat date in SubDate and LatestCollectedDate using formula format')
    apply_date_formula(subdate_col_idx, new_df["SubDate"])
    apply_date_formula(latest_collected_col_idx, new_df["LatestCollectedDate"])

    logging.info(f'Save {file_name} in {folder_path}')
    wb.save(output_file_path)

    # table info
    count_onetime = (new_df['Frequency'] == 0).sum()
    count_pledge = (new_df['Frequency'] > 0).sum()

    logging.info(f'Table length: {len(new_df)}, Total Pledge: {count_pledge}, Total One-Time: {count_onetime}')
    logging.info(f'Process 1 completed')

    

    
    
def extract_batchid_and_create_filename(file):
    match = re.search(r'\b\d{6}\b', file)
    return f'Extracted result from {match.group()}.xlsx'


def extract_data_from_result_file(folder_path, file):
    """
    What does this function do?
    1. Open .csv file from file_path given.
    2. Parse the data into list and convert to dataframe and split data into 2 column by '='
    3. Return as dataframe
    """

    logging.info('Process 2: Extract data from EBC result file')
    logging.debug('Create file path')
    file_path = os.path.join(folder_path, file)

    filename = file

    logging.debug('Create dictionary with desired key to parse')
    desired_keys = {'merchantReferenceCode',
                    'paySubscriptionCreateReply_subscriptionID',
                    'paySubscriptionCreateReply_instrumentIdentifierID'
                    }

    logging.debug('Create list for parsed data')
    parsed_data = []

    
    logging.info(f'Read file: {file}')
    logging.info('Parsed data from result file and create dataframe table')
    with open(file_path, 'r') as file:
        lines = file.readlines()
        
        # Skip the first two rows 
        if len(lines) > 2:
            lines = lines[2:]

        logging.debug('Parsed data from file')
        for line in lines:
            line = line.strip()
            if line:  # Skip empty lines
                items = line.split(',')
                parsed_dict = {key: '' for key in desired_keys}
                for item in items:
                    if '=' in item:
                        key, value = item.split('=', 1)  # Split only on the first '='
                        key = key.strip()
                        value = value.strip()
                        if key in desired_keys:
                            parsed_dict[key] = value
                parsed_data.append(parsed_dict)
    
    logging.debug('Turn parsed data into table and save in variable')
    parsed_df = pd.DataFrame(parsed_data)

    logging.debug('Create new file name')
    new_file_name = extract_batchid_and_create_filename(filename)

    logging.info(f'Save {new_file_name} in {folder_path}')
    parsed_df.to_excel(os.path.join(folder_path, new_file_name), index=False)

    # table info
    table_length = len(parsed_df)
    token_count = len(parsed_df['paySubscriptionCreateReply_subscriptionID'])
    failed_token_count = (
        (parsed_df['paySubscriptionCreateReply_subscriptionID'] == '') | 
        (parsed_df['paySubscriptionCreateReply_subscriptionID'].isna())
        ).sum()

    logging.info(f'Table length: {table_length}, Token Success: {token_count}, Token Failed: {failed_token_count} ')
    logging.info('Process 2 completed')
    

def process_sw_and_result_file(folder_path):
    #folder_path = r'C:\Users\mfmohammad\OneDrive - UNICEF\Documents\2 Areas\Saleswork Process\May\20250529'
    
    logging.info('Process Starts!')
    start_time = time.time() 

    directory = os.listdir(folder_path)

    for file in directory:
        if 'ExportDailyDataSG' in file or 'ExportDailyData' in file:

            convert_sw_file(folder_path, file)

        elif 'unicef_malaysia' in file and 'reply.all' in file:
            
            extract_data_from_result_file(folder_path, file)

    elapsed_time = time.time() - start_time
    logging.info(f'Process Completed. Elapsed time : {elapsed_time:.2f} seconds.')

"""       
if __name__ == '__main__' :
    process_sw_and_result_file()
"""